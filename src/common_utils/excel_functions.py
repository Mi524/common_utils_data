from openpyxl import Workbook as openpyxl_workbook
from openpyxl.styles import Color,PatternFill,Side, Font, Border,Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
from xlsxwriter.exceptions import FileCreateError
from xlsxwriter import Workbook
from common_utils.os_functions import check_create_new_folder
from common_utils.sequence_functions import duplicate_elem_add_seq
from decimal import Decimal
import datetime 
import re 
import os 
import math 
import warnings 
import numpy as np
import xlwings as xw 

from pandas.core.indexes.multi import MultiIndex

"""pandas save_excel """
def pd_save_excel(df,save_path,startrow=0,startcol=0):

	close_tag = 0
	while close_tag <= 0:
		try:
			df.to_excel(save_path,index=False,startrow=startrow,startcol=startcol)
			close_tag += 1 
		except (PermissionError,FileCreateError) as e :
			input('\nFailed to write file!\n  Please Close "{}" Then Press Enter to Continue'.format(save_path))
	print('{0} Saved'.format(save_path))

"""这里主要是xlsxwriter的方法"""
def autofit_column_width(xlsxwriter_ws,header_columns,content_columns=[],wrap_text=False, **kwargs):
	"""根据表头的字符长度自动调整EXCEL表头的列宽
	   content_columns是专门给特长的列加宽
	"""
	min_column_width = kwargs.get('min_column_width',10)

	cn_pat = "[\u4e00-\u9fa5]+"
	en_pat = "[^\u4e00-\u9fa5]+"

	#先统一把表头转成文字格式
	header_columns = [str(x) for x in header_columns]

	if wrap_text == False:
		length_list = [ len(''.join(re.findall(cn_pat,c)))*2 + len(''.join(re.findall(en_pat,c)))*1  + 4 \
							for c in list(header_columns)] 
		for i, width in enumerate(length_list):
			xlsxwriter_ws.set_column(i,i,width)
		#防止最后一列没修改成功
		xlsxwriter_ws.set_column(i+1,i+1,width)

	else: #如果采用wrap_text方式，计算两层的中文wrap结果 
		length_list = [ math.ceil(len(''.join(re.findall(cn_pat,c)))/2) * 2 + \
						math.ceil(len(''.join(re.findall(en_pat,c)))/2) * 1 + 2 \
							for c in list(header_columns)] 

		#确保每列的宽度至少是4 
		length_list = [x if x >= min_column_width else min_column_width for x in length_list ]

		for i, width in enumerate(length_list):
			xlsxwriter_ws.set_column(i,i,width)
		#防止最后一列没修改成功
		xlsxwriter_ws.set_column(i+1,i+1,width)

	#给特别的内容列加宽
	if content_columns:
		for c in content_columns:
			c_index = header_columns.index(c)
			xlsxwriter_ws.set_column(c_index,c_index,80)

def adjust_pct_decimal_format(original_number_format,data):
	#如果某一行/列的数据数字50%及以上的部分到第二个百分点无法区分上升或下降,比如0.001%, 0.002%
	#保留2个小数点,四舍五入都是约等于0.00%,则往下一级去扩展保留更多的小数点位数
	original_decimal = original_number_format[:-1].split('.')

	if original_decimal :
		original_decimal_places = len(original_decimal[1]) 
	else:
		original_decimal_places = 2

	#找到该列数据小数点最长的数字的小数点位数
	biggest_decimal_places = max([abs(Decimal(str(x)).as_tuple().exponent) for x in data if type(x)==float ])

	#当约到某一位的数字超过50%时，加一个小数位
	data_decimal_check = [ round(x,original_decimal_places+2) for x in data if type(x)==float ]

	#众数的数量如果超过50%
	mode_counter = Counter(data_decimal_check).most_common()[0]
	mode = mode_counter[0]
	mode_counter = mode_counter[1]
	
	mode_decimal_places = abs(Decimal(mode).as_tuple().exponent)

	#不能把original_decimal_format直接赋值给next_decimal_format,不可修改
	next_decimal_format = original_number_format

	if  biggest_decimal_places > original_decimal_places and mode_counter/len(data_decimal_check) >= 0.5:
		next_decimal_format = '0' * (original_decimal_places + 1)
		next_decimal_format = '0.{}%'.format(next_decimal_format)

		next_decimal_format = adjust_pct_decimal_format(next_decimal_format,data)

	return next_decimal_format

def write_row_format(xlsxwriter_ws,data,start_pos,cell_format={},direction=0):
	"""fit in row or col parameter to write_row method according to direction param
	:param row : start row zero indexed 
	:param col : start col zero indexed
	:param pct_format : how to display num_format
	:param direction : 1/horizontal ，0/vertical
	"""
	data = [ x if x == x else '' for x in data]
	if direction == 1:
		xlsxwriter_ws.write_column(row=0,col=start_pos,data=data,cell_format=cell_format)
	else :
		xlsxwriter_ws.write_row(row=start_pos,col=0,data=data,cell_format=cell_format)

def write_multi_tables(xlsxwriter_wb,sheet_name,df_list,direction=0,**kwargs):
	"""将多个小型数据表写入一个sheet里面
	   0/horizontal 从上往下顺序写入，1/vertical 从左往右顺序写入
	   有额外的kwargs参数，方便输入遇到哪个字符串就用百分比格式
	"""
	percent_str = kwargs.get('percent_str','占比')
	auto_adjust_pct =  kwargs.get('auto_adjust_pct',False)
	normal_format =  kwargs.get('normal_format',{'font_name':'calibri','font_size':11})
	date_format = kwargs.get('date_format','%Y-%m-%d')

	#通用表头格式
	background_color = rgb_convert_hex([220,230,241])
	border_color = rgb_convert_hex([149,179,215])

	header_format = xlsxwriter_wb.add_format({'font_name':'微软雅黑','bold':False,'font_size':11,
											  'bg_color':background_color })
	header_format.set_bottom()
	header_format.set_bottom_color(border_color)

	normal_format = xlsxwriter_wb.add_format(normal_format)

	xlsxwriter_ws = xlsxwriter_wb.add_worksheet(sheet_name)

	#记录表头
	horizontal_columns = [ ]
	vertical_columns = [ ]
	#记录每次位置
	position_counter = 0

	for df in df_list:
		#如果其中某个或所有index不是常规的数字, reset_index
		if not all([type(x) == int for x in list(df.index) ]) :
			df =  df.reset_index()

		#记录横着写的表格所有表头
		horizontal_columns += list(df.columns) + [' ' * 8]
		vertical_columns = list(df.columns)

		#先判断这个表格是否要按照百分比格式写入
		if percent_str in df.columns.values[0]: 
			pct_tag = True
		else: 
			pct_tag = False 

		#表头reset
		df = df.T.reset_index().T.reset_index(drop=True)
		#如果是以列的方式写入
		if direction == 1:
			data_set = [ (i,list(df[x])) for i,x in enumerate(df.columns)]
		else:
			data_set = df.iterrows()

		is_header = True
		for i,data in data_set:
			if is_header:
				#写入表头
				header = list(df.columns.values)
				write_row_format(xlsxwriter_ws,data=data,start_pos=position_counter,
								cell_format=header_format,direction=direction)
				position_counter += 1
				is_header = False 

			else:
				# 如果是日期，转成对应的日期格式
				data = [ datetime.datetime.strftime(x,date_format) if isinstance(x, datetime.datetime) else x for x in data ]
				data = [ float(x) if type(x) != str else x for x in data ]

				if pct_tag:
					#百分比需要判断放入多少个小数点的百分比,不在这里加入
					number_format = kwargs.get('num_format','0.00%')
					if auto_adjust_pct :  #
						number_format = adjust_pct_decimal_format(number_format,data)

					pct_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':number_format})

					write_row_format(xlsxwriter_ws,data=data,start_pos=position_counter,
									 cell_format=pct_format,direction=direction)
				else:
					write_row_format(xlsxwriter_ws,data=data,start_pos=position_counter,
									cell_format=normal_format, direction=direction)

				position_counter += 1 

		#不同表格换行
		position_counter += 1

		if direction == 1:
			autofit_column_width(xlsxwriter_ws,horizontal_columns)
		else:
			autofit_column_width(xlsxwriter_ws,vertical_columns)


def df_sheet_check(df_list,sheet_name_list):
	#防止填入的不是列表格式,防止填入的sheet_name 出现重复
	if type(df_list) != list:
		df_list = [df_list]
	if type(sheet_name_list) != list:
		sheet_name_list = [sheet_name_list]
	
	if len(df_list) != len(sheet_name_list):
		print('写入的数据表数量和填入的Sheet名称数量不相等,将使用默认Sheet名称')

	sheet_name_list = duplicate_elem_add_seq(sheet_name_list)

	if sheet_name_list == None or sheet_name_list == [ ] :
		sheet_name_list= []
		for i in range(len(df_list)):
			sheet_name_list.append('Sheet {}'.format(i+1))

	return df_list, sheet_name_list
	
def write_pct_columns(save_path,df_list,sheet_name_list=[],pct_columns=[],content_columns=[],**kwargs):
	"""给定一个表格，用百分比写入包含指定文字的列,表头wrap_text, 添加选项"""

	xlsxwriter_wb = Workbook(save_path)

	num_format = kwargs.get('num_format','0.00%')
	#通用表头格式
	background_color = rgb_convert_hex([220,230,241])
	border_color = rgb_convert_hex([149,179,215])
	header_format = xlsxwriter_wb.add_format({'font_name':'微软雅黑','bold':False,'font_size':11,
											  'bg_color':background_color })
	header_format.set_bottom()
	header_format.set_bottom_color(border_color)

	header_format.set_text_wrap()
	header_format.set_center_across()

	date_format =  xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'yyyy/mm/dd'})
	pct_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'0.00%'})
	text_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'@'})
	normal_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11})

	df_list, sheet_name_list = df_sheet_check(df_list, sheet_name_list)

	for df,sheet_name in zip(df_list,sheet_name_list):
		xlsxwriter_ws = xlsxwriter_wb.add_worksheet(sheet_name)

		header_columns = list(df.columns)

		if not df.empty:
			pct_columns = [ x for x in header_columns if [y for y in pct_columns if y in x ]]
			#写入表头
			xlsxwriter_ws.write_row(row=0,col=0,data=header_columns,cell_format=header_format)

			for column_index,column in enumerate(header_columns):
				column_values = df[column].to_list()

				if column in pct_columns:  #如果属于百分比列 
					xlsxwriter_ws.write_column(row=1,col=column_index,data=column_values,cell_format=pct_format)
				#如果本列是日期类型,采用日期格式写入, 第二步应该到判断content_columns 还是 date? -- 以可以输入的参数为优先做判断
				elif np.issubdtype(df[column].dtype, np.datetime64) :
					column_format = date_format
				else:
					column_values = [ x if x == x else '' for x in column_values  ]
					column_format =  normal_format

				xlsxwriter_ws.write_column(row=1,col=column_index,data=column_values,cell_format=current_format)

			#表头加上选项
			xlsxwriter_ws.autofilter(0,0,len(df)-1,len(header_columns)-1)

			autofit_column_width(xlsxwriter_ws,header_columns,content_columns=content_columns,wrap_text=True)

			#冻结第一行窗口
			xlsxwriter_ws.freeze_panes(1,0)
		else: #如果没有数据，默认写入表头即可
			xlsxwriter_ws.write_row(row=0,col=0,data=header_columns,cell_format=header_format)

	save_xlsxwriter_wb(xlsxwriter_wb,save_path)

def replace_invalid_strs(string):
	invalid_strs = [ '\ud835' ]
	for i in invalid_strs:
		string = string.replace(i,'')
	return string 

def refresh_excel_calculations(file_path):
	if not os.path.isabs(file_path):
		file_path = os.path.join(os.getcwd(), file_path)

	xlapp = xw.App(visible=False)
	xlapp.display_alerts = False
	xlwb = xw.Book(file_path)
	xlwb.Visible = False
	xlapp.calculate()
	xlwb.save()
	xlapp.quit()

	del xlapp

# def refresh_excel_calculations(file_path):
# 	file_path = os.path.join(os.getcwd(), file_path )
# 	#检查输出文件夹 是否存在
# 	if not os.path.isfile(file_path):
# 		enter_exit("File not found: ",file_path)

# 	#打开文档，运行VBA, 如果有报错，everything 搜索gen_py文件 删除，让系统重新生成新的gen_py
# 	xlapp = client.gencache.EnsureDispatch('Excel.Application')   
# 	xlapp.Visible = 0  
# 	xlapp.DisplayAlerts = False 
# 	xlwb = xlapp.Workbooks.Open(file_path) 
# 	xlwb.Visible = 0
# 	xlapp.Calculate()
# 	xlwb.Close(True)
# 	xlapp.Quit

def write_format_columns(save_path,df_list,sheet_name_list=[],pct_columns=[],round_columns=[],content_columns=[],**kwargs):
	"""给定一个表格，用百分比写入包含指定文字的列,表头wrap_text, 添加选项,上面write_pct_columns方式不写是不希望每列str都变成text格式，
	   正常应该是常规格式
	   基本和前面的write_pct_columns相同，区别在于加上了整列如果都是str用text格式写入的方式"""
	xlsxwriter_wb = Workbook(save_path)

	min_column_width = kwargs.get('min_column_width', 12)

	round_list = kwargs.get('round_list',[ '标准差','平均','方差'])
	pct_list = kwargs.get('pct_list',['rate','百分比','占比','比例','比率'])

	#通用表头格式
	background_color = rgb_convert_hex([220,230,241])
	border_color = rgb_convert_hex([149,179,215])
	header_format = xlsxwriter_wb.add_format({'font_name':'微软雅黑','bold':False,'font_size':11,
											  'bg_color':background_color })
	header_format.set_bottom()
	header_format.set_bottom_color(border_color)

	header_format.set_text_wrap()
	header_format.set_center_across()

	date_format =  xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'yyyy/mm/dd'})
	pct_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'0.00%'})
	text_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'@'})
	normal_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11})
	round_format = xlsxwriter_wb.add_format({'font_name':'calibri','font_size':11,'num_format':'0.00'})
	
	df_list, sheet_name_list = df_sheet_check(df_list, sheet_name_list)

	for df,sheet_name in zip(df_list,sheet_name_list):
		xlsxwriter_ws = xlsxwriter_wb.add_worksheet(sheet_name)

		#是否属于MultiIndex,如果是，需要确保写入完整--暂时不添加该功能
		# if isinstance(df.column, MultiIndex):

		header_columns = list(df.columns)

		if not df.empty:
			pct_columns = [ x for x in header_columns if [y for y in pct_columns if y in x ]]
			#写入表头
			xlsxwriter_ws.write_row(row=0,col=0,data=header_columns,cell_format=header_format)

			for column_index,column in enumerate(header_columns):
				column_values = df[column].to_list()
				if column in pct_columns or len([ x for x in pct_list if x in column]) > 0:  #如果属于百分比列 
					column_format = pct_format
				elif column in round_columns or len([ x for x in round_list if x in column]) > 0  : #如果属于需要round的列
					column_format = round_format
				#判断是否整列都属于str格式，采用text格式写入
				elif df[column].apply(type).eq(str).all():
					column_format = text_format
				#如果本列是日期类型,采用日期格式写入, 第二步应该到判断content_columns 还是 date? -- 以可以输入的参数为优先做判断
				elif np.issubdtype(df[column].dtype, np.datetime64)  :
					column_values = [ x if x == x else '' for x in column_values]
					column_format = date_format
				#是否整列都是float格式
				# elif df[column].apply(type).eq(float).all():
				# 	column_format = round_format
				else:
					column_values = [ x if x == x and type(x) == str else x for x in column_values  ]
					column_values = [ x if x == x else '' for x in column_values]
					column_format =  normal_format

				xlsxwriter_ws.write_column(row=1,col=column_index,data=column_values,cell_format=column_format)

			#表头加上选项
			xlsxwriter_ws.autofilter(0,0,len(df)-1,len(header_columns)-1)

			autofit_column_width(xlsxwriter_ws,header_columns,content_columns=content_columns,
											wrap_text=True, min_column_width=min_column_width)

			#冻结第一行窗口
			xlsxwriter_ws.freeze_panes(1,0)
		else: #如果没有数据，默认写入表头即可
			xlsxwriter_ws.write_row(row=0,col=0,data=header_columns,cell_format=header_format)

	save_xlsxwriter_wb(xlsxwriter_wb,save_path)

"""以下部分主要是openpyxl的方法"""
def rgb_convert_hex(rgb_list):
	#防止填入了超过255或者小于0的数值
	rgb_list = [ max(0,min(int(x),255)) for x in rgb_list]
	hex_value = '{0:02x}{1:02x}{2:02x}'.format(rgb_list[0],rgb_list[1],rgb_list[2]).upper()
	return hex_value


def get_column_row_num(worksheet):
	"""
	calculate the number of columns and rows from a worksheet 
	:param worksheet: openpyxl worksheet
	:return : column_letters, row_numbers 
	"""
	#计算一共有多少列
	column_num = len(list(worksheet['1']))
	column_letters = [get_column_letter(x).upper() for x in range(1,column_num+1)]

	#计算一共有多少行
	row_num = len(list(worksheet['A']))
	row_numbers = [ x for x in range(1,row_num+1) ]

	return column_letters,row_numbers

def format_custom_cells(worksheet,custom_cells_dict,custom_cell_color):
	"""
	处理特殊的单元格
	:param custom_cells_dict :默认key代表要修改的sheet_name,value是list类型，第一个元素是列，第二个元素的行
	:custom_cell_color: 指定单元格的字体颜色
	"""
	custom_cell_color = rgb_convert_hex(custom_cell_color)
	custom_font = Font(color=custom_cell_color)

	#需要调整的sheet_name
	#检查传入的sheet是否是需要修改的匹配的sheet_name
	for c_key in custom_cells_dict:
		if c_key == worksheet.title:
			#如果是这个需要修改的sheet 
			#第一个元素为列，第二个元素代表行
			c_column_list = custom_cells_dict[c_key][0]
			c_row_list = custom_cells_dict[c_key][1]

			#防止填入的数字有0
			if 0 in c_column_list:
				c_column_list = [x+1 for x in c_column_list]		
			if 0 in c_row_list:
				c_row_list = [x+1 for x in c_row_list]

			#循环修改格式
			for c_column in c_column_list:
				c_column_letter = get_column_letter(c_column)
				for c_row in c_row_list:
					c_target = c_column_letter + str(c_row)
					worksheet[c_target].font = custom_font

def format_table(worksheet,format_style,kwargs):
	"""
	传入worksheet进行统一格式调整后返回worksheet对象
	:param worksheet :openpyxl worksheet 
	:param format_style : 1/text or 2/data ,default is None
	:param wrap_text : the column letter you want to have auto line break, default is None
	:header_color : rgb value in list format, None is transparent, 'default' is light blue
	"""
	if kwargs == {}:
		kwargs = defaultdict()

	#表头颜色，最后一行字体是否加粗
	header_color = kwargs.get('header_color','default')
	last_row_bold = kwargs.get('last_row_bold',False)
	#第一行第一列的高度和宽度
	first_row_height = kwargs.get('first_row_height',27)
	first_column_width = kwargs.get('first_column_width',20)
	#主体部分宽和高
	main_part_width = kwargs.get('main_part_width',14)
	main_part_height = kwargs.get('main_part_height',14)
	#定制单元格
	custom_cells_dict = kwargs.get('custom_cells_dict',None)
	custom_cell_color = kwargs.get('custom_cell_color',[255,0,0]) #red 

	column_letters,row_numbers = get_column_row_num(worksheet)

	# =========== 通用的格式部分 ==========
	#第一行设置格式 _ 自动换行 _ 居中,是否旋转文字 等条件 
	row_1_alignment = Alignment(    horizontal = 'center',
									vertical = 'center',
									text_rotation=0,
									wrap_text=True,
									shrink_to_fit=False,
									indent=0  )
	#字体加粗
	row_1_font = Font(bold=True)

	#循环第一行格式
	for cell_1 in worksheet['1']:
		cell_1.alignment = row_1_alignment
		cell_1.font = row_1_font

	#表头颜色
	default_header_color = [189,215,238]  #浅蓝色
	
	if header_color == 'default':
		header_color = default_header_color

	if header_color != None:
		#把rgb的颜色格式转成aRGB hex value 
		header_color = rgb_convert_hex(header_color)
		#循环第一行背景颜色填充
		row_1_colorFill = PatternFill(start_color = header_color,
										end_color = header_color,
										fill_type ='solid')
		for cell_1 in worksheet['1']:
			cell_1.fill = row_1_colorFill

	#默认全部全部表格居左
	cell_alignment = Alignment(horizontal = 'left',vertical = 'center')

	for row in list(worksheet.rows)[1:]:
		for cell in row:
			cell.alignment = cell_alignment

	#循环调整行宽（默认行宽太窄）,以下代码容易出现错误 且不会提示哪里错，行对应是高，列对应是输入宽度
	#wrap_text很容易受别的alignment影响，在text格式下也不再进行调整
	for r in row_numbers:
		worksheet.row_dimensions[r].height = main_part_height
	#调整行高
	for c in column_letters:
		worksheet.column_dimensions[c].width = main_part_width

	#是否包含需要特殊格式的单元格
	if custom_cells_dict != None:
		format_custom_cells(worksheet,custom_cells_dict,custom_cell_color)

	# ========  1 TEXT .普通文字型表格 : 最后一列调整成特别的宽度============
	if str(format_style).lower() == 'text' or format_style == 1:
		worksheet.column_dimensions[column_letters[-1]].width = 80

		#wrap_text会被固定的行高限定住，如果想auto linebreak还是需要计算单元格里面的字体内容，通过内容和宽度得到应该调整的行高，
		#当前虽然没办法自动换行，但是还是需要添加这个条件，否则打开excel选择自动换行也不生效，还要选个行高自动调整
		wrap_text_alignment = Alignment(wrap_text=True)
		for cell in list(worksheet[column_letters[-1]])[1:]:
			cell.alignment = wrap_text_alignment

	# ============= 2 DATA 数据型表格 ==============
	if str(format_style).lower() == 'data' or format_style == 2:
		#第一行加高和第一列加宽
		worksheet.row_dimensions[1].height = first_row_height
		worksheet.column_dimensions['A'].width = first_column_width

		#循环第一列居左
		for cell_A in worksheet['A']:
			cell_A.alignment = Alignment(horizontal='left',vertical='center')

		#循环调整其他部分的全部列宽
		for c in column_letters[1:]:
			worksheet.column_dimensions[c].width = main_part_width
		#循环调整其他部分的全部行高
		for r in row_numbers[1:]:
			worksheet.row_dimensions[r].height = main_part_height

		#循环数字部分主体靠右展示
		for row in list(worksheet.rows)[1:]:
			for cell in list(row)[1:]:
				cell.alignment = Alignment(horizontal='right',vertical='center')

		#最后一行是否要加粗
		if last_row_bold == True:
			for cell in list(worksheet.rows)[-1]:
				cell.font = Font(bold=True)
				cell.border = Border(bottom=Side(border_style="thin", color="000000"))

	return worksheet 


def save_xlsxwriter_wb(xlsxwriter_wb,save_path):
	#检查是否存在save_path路径
	check_create_new_folder(save_path)
	close_tag = 0
	while close_tag <= 0:
		try:
			xlsxwriter_wb.close()
			close_tag += 1 
		except FileCreateError:
			input('\nFailed to write file!\n  Please Close "{}" Then Press Enter to Continue'.format(save_path))
	print('{0} Saved'.format(save_path))


def save_csv(df, save_path):
	check_create_new_folder(save_path)
	close_tag = 0
	while close_tag <= 0:
		try:
			df.to_csv(save_path,index=False)
			close_tag += 1 
		except PermissionError:
			input('\nFailed to write file!\n  Please Close "{}" Then Press Enter to Continue'.format(save_path))
	print('{0} Saved'.format(save_path))


def save_excel(df_list,save_path,sheet_name_list=None,format_style=None,**kwargs):
	"""
	通过openpyxl写入pandas表格，可以修改写入格式
	:param df_list : df list need to write 
	:param save_path : target path
	:param sheet_name_list : sheet_names coresponding to df_list
	:param format_style : default is None, bolded header;  text or 1  (change header format 
	 and column width only); 'data' or 2 adjust the header and the first column
	"""
	wb = openpyxl_workbook()

	#防止填入的不是列表格式
	if type(df_list) != list:
		df_list = [df_list]
	if type(sheet_name_list) != list:
		sheet_name_list = [sheet_name_list]
	
	if len(df_list) != len(sheet_name_list):
		print('写入的数据表数量和填入的Sheet名称数量不相等,将使用默认Sheet名称')
		
	if sheet_name_list == [None] :
		sheet_name_list= []
		for i in range(len(df_list)):
			sheet_name_list.append('Sheet {}'.format(i+1))
	
	for df, sheet_name in zip(df_list,sheet_name_list):
		if not df.empty:
			ws = wb.create_sheet(title=sheet_name)
			#写入表头
			column_list = df.columns.tolist()
			ws.append(column_list)
			counter = 0 
			for value in df.values:
				value = [ x.replace('\x01','') if type(x) == str  else x for x in value ]
				counter += 1 
				try:
					ws.append(list(value))
				except:
					print(f'本行数据出错：{sheet_name},{counter}')
					print(value)
			ws = format_table(ws,format_style=format_style,kwargs=kwargs)

	#删除默认的sheet
	del wb['Sheet']

	#保存全部文档
	check_close = 0  
	while check_close <= 0:
		try:
			wb.save(save_path)
			check_close += 1 
		except PermissionError:
			input('无法正常记录，请关闭{}后摁回车键继续'.format(save_path))

	print('{0} 已保存'.format(save_path))


